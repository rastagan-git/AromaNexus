"""Benchmark XLSX checkpoint serialization without provider or network work.

Run from the repository root after installing the development environment::

    python -m benchmarks.checkpoint_serialization --runs 3

The benchmark builds one synthetic workbook, skips every input row before a
PubChem client can be called, and compares median end-to-end runtime with
checkpoints disabled and with a 25-row interval.
"""

from __future__ import annotations

import argparse
import statistics
import tempfile
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from aromanexus.workflows import run_pubchem


class _NoNetworkPubChem:
    def __init__(self) -> None:
        self.calls = 0

    def lookup(self, identifier: object, *, include_odor: bool = True):
        self.calls += 1
        raise AssertionError(f"Benchmark unexpectedly called PubChem for {identifier!r}")


def _silent_progress(*_args: object) -> None:
    return None


def _create_fixture(path: Path, rows: int) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["Name", "Measured value", "Calculated value", "Existing CAS"])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for position in range(1, rows + 1):
        excel_row = position + 1
        worksheet.append(
            [
                f"synthetic-compound-{position:04d}",
                position / 10,
                f"=B{excel_row}*2",
                "",
            ]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:D{rows + 1}"
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 16

    notes = workbook.create_sheet("Notes")
    notes.append(["Benchmark", "No-network checkpoint serialization"])
    notes.append(["Rows", rows])
    workbook.save(path)
    workbook.close()


def _time_run(
    source: Path,
    directory: Path,
    *,
    checkpoint_every: int,
    sequence: int,
    rows: int,
) -> float:
    destination = directory / f"result-{checkpoint_every}-{sequence}.xlsx"
    client = _NoNetworkPubChem()
    started = time.perf_counter()
    summary = run_pubchem(
        source,
        client,
        output_path=destination,
        sheet_name="Data",
        identifier_column="Name",
        skip_patterns=(r"^.*$",),
        include_odor=False,
        include_provenance=True,
        checkpoint_every=checkpoint_every,
        progress=_silent_progress,
    )
    elapsed = time.perf_counter() - started

    if client.calls:
        raise AssertionError(f"Expected zero provider calls, observed {client.calls}")
    if summary.status_counts != {"skipped": rows}:
        raise AssertionError(f"Unexpected status counts: {summary.status_counts}")
    if not destination.is_file():
        raise AssertionError(f"Benchmark output was not created: {destination}")
    partial = destination.with_name(f"{destination.stem}.partial{destination.suffix}")
    if partial.exists():
        raise AssertionError(f"Successful run left a checkpoint behind: {partial}")

    destination.unlink()
    return elapsed


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--rows", type=int, default=262, help="Synthetic input rows (default: 262)")
    parser.add_argument("--runs", type=int, default=3, help="Measured runs per mode (default: 3)")
    parser.add_argument("--warmups", type=int, default=1, help="Warm-up runs per mode (default: 1)")
    parser.add_argument(
        "--checkpoint-every",
        type=int,
        default=25,
        metavar="N",
        help="Enabled checkpoint interval compared with 0 (default: 25)",
    )
    parser.add_argument(
        "--max-ratio",
        type=float,
        help="Return a failing exit code if enabled/disabled median exceeds this ratio",
    )
    args = parser.parse_args()
    if args.rows < 1:
        parser.error("--rows must be at least 1")
    if args.runs < 1:
        parser.error("--runs must be at least 1")
    if args.warmups < 0:
        parser.error("--warmups cannot be negative")
    if args.checkpoint_every < 1:
        parser.error("--checkpoint-every must be at least 1")
    if args.max_ratio is not None and args.max_ratio <= 0:
        parser.error("--max-ratio must be positive")
    return args


def main() -> int:
    args = _parse_args()
    samples: dict[int, list[float]] = {0: [], args.checkpoint_every: []}

    with tempfile.TemporaryDirectory(prefix="aromanexus-checkpoint-benchmark-") as temporary:
        directory = Path(temporary)
        source = directory / "synthetic-input.xlsx"
        _create_fixture(source, args.rows)

        total_rounds = args.warmups + args.runs
        sequence = 0
        for round_number in range(total_rounds):
            modes = (0, args.checkpoint_every)
            if round_number % 2:
                modes = tuple(reversed(modes))
            for checkpoint_every in modes:
                sequence += 1
                elapsed = _time_run(
                    source,
                    directory,
                    checkpoint_every=checkpoint_every,
                    sequence=sequence,
                    rows=args.rows,
                )
                if round_number >= args.warmups:
                    samples[checkpoint_every].append(elapsed)

    disabled_median = statistics.median(samples[0])
    enabled_median = statistics.median(samples[args.checkpoint_every])
    ratio = enabled_median / disabled_median
    overhead = (ratio - 1) * 100

    print(
        f"rows={args.rows} measured_runs={args.runs} warmups={args.warmups} "
        f"checkpoint_interval={args.checkpoint_every}"
    )
    print(f"checkpoint=0 median:  {disabled_median:.3f}s  samples={samples[0]}")
    print(
        f"checkpoint={args.checkpoint_every} median: "
        f"{enabled_median:.3f}s  samples={samples[args.checkpoint_every]}"
    )
    print(f"enabled/disabled ratio: {ratio:.3f} ({overhead:+.1f}% overhead)")

    if args.max_ratio is not None and ratio > args.max_ratio:
        print(f"FAIL: ratio {ratio:.3f} exceeds --max-ratio {args.max_ratio:.3f}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
