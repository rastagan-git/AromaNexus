"""Inspect an AromaNexus workbook without modifying it."""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[4]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from aromanexus.excel_io import (  # noqa: E402
    FORMULA_PREFIXES,
    read_table,
    workbook_sheet_names,
)
from aromanexus.identifiers import is_valid_cas, normalize_cas  # noqa: E402


def _stable_cell_value(value: object) -> str | None:
    if value is None:
        return None
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def _sheet_summary(worksheet, cached_worksheet) -> dict[str, object]:
    digest = hashlib.sha256()
    formula_cache_digest = hashlib.sha256()
    formula_cells = 0
    cached_formula_results = 0
    styled_cells = 0
    nonempty_cells = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.has_style:
                styled_cells += 1
            if cell.value is None:
                continue
            nonempty_cells += 1
            formula_cells += int(cell.data_type == "f")
            value = _stable_cell_value(cell.value)
            record = json.dumps(
                [cell.coordinate, cell.data_type, value],
                ensure_ascii=False,
                separators=(",", ":"),
            )
            digest.update(record.encode("utf-8"))
            digest.update(b"\n")
            if cell.data_type == "f":
                cached_cell = cached_worksheet[cell.coordinate]
                cached_value = _stable_cell_value(cached_cell.value)
                cached_formula_results += int(cached_value is not None)
                cached_record = json.dumps(
                    [cell.coordinate, cached_cell.data_type, cached_value],
                    ensure_ascii=False,
                    separators=(",", ":"),
                )
                formula_cache_digest.update(cached_record.encode("utf-8"))
                formula_cache_digest.update(b"\n")
    return {
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "nonempty_cells": nonempty_cells,
        "formula_cells": formula_cells,
        "cached_formula_results": cached_formula_results,
        "formula_cache_sha256": formula_cache_digest.hexdigest(),
        "styled_cells": styled_cells,
        "content_sha256": digest.hexdigest(),
        "row_heights": {
            str(index): dimension.height
            for index, dimension in worksheet.row_dimensions.items()
            if dimension.height is not None
        },
        "column_widths": {
            name: dimension.width
            for name, dimension in worksheet.column_dimensions.items()
            if dimension.width is not None
        },
        "freeze_panes": str(worksheet.freeze_panes or ""),
        "auto_filter": str(worksheet.auto_filter.ref or ""),
        "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
        "tables": list(worksheet.tables),
        "data_validations": len(worksheet.data_validations.dataValidation),
        "conditional_formatting_ranges": len(worksheet.conditional_formatting),
    }


def inspect(
    path: Path,
    cas_column: str | None = None,
    sheet_name: str | None = None,
) -> dict[str, object]:
    frame = read_table(path, sheet_name=sheet_name)
    detected_cas = cas_column or next(
        (name for name in ("CAS Number", "CAS", "cas", "cas_number") if name in frame.columns),
        None,
    )
    if detected_cas is not None and detected_cas not in frame.columns:
        available = ", ".join(map(str, frame.columns))
        raise ValueError(f"CAS column {detected_cas!r} was not found. Available: {available}")
    formula_like = 0
    for column in frame.columns:
        formula_like += int(
            frame[column]
            .map(lambda value: isinstance(value, str) and value.startswith(FORMULA_PREFIXES))
            .sum()
        )
    report: dict[str, object] = {
        "path": str(path.resolve()),
        "rows": len(frame),
        "columns": [str(column) for column in frame.columns],
        "duplicate_rows": int(frame.duplicated().sum()),
        "formula_like_cells": formula_like,
        "missing_by_column": {
            str(column): int(frame[column].isna().sum() + frame[column].eq("").sum())
            for column in frame.columns
        },
    }
    if path.suffix.lower() == ".xlsx":
        sheets = workbook_sheet_names(path)
        selected_sheet = sheet_name or sheets[0]
        workbook = load_workbook(
            path,
            read_only=False,
            data_only=False,
            keep_links=True,
            rich_text=True,
        )
        cached_workbook = load_workbook(
            path,
            read_only=False,
            data_only=True,
            keep_links=True,
            rich_text=True,
        )
        try:
            sheet_summaries = {
                worksheet.title: _sheet_summary(
                    worksheet,
                    cached_workbook[worksheet.title],
                )
                for worksheet in workbook.worksheets
            }
            report["workbook"] = {
                "title": workbook.properties.title or "",
                "creator": workbook.properties.creator or "",
                "sheets": sheets,
                "selected_sheet": selected_sheet,
                "sheet_summaries": sheet_summaries,
                "selected_sheet_summary": sheet_summaries[selected_sheet],
            }
        finally:
            workbook.close()
            cached_workbook.close()
    if detected_cas:
        normalized = frame[detected_cas].map(normalize_cas)
        valid_mask = normalized.map(is_valid_cas)
        nonempty_mask = normalized.ne("")
        invalid = normalized[nonempty_mask & ~valid_mask]
        report["cas"] = {
            "column": detected_cas,
            "nonempty": int(nonempty_mask.sum()),
            "valid": int(valid_mask.sum()),
            "invalid": int((nonempty_mask & ~valid_mask).sum()),
            "invalid_examples": list(dict.fromkeys(invalid.astype(str)))[:10],
            "duplicates": int(normalized[nonempty_mask].duplicated().sum()),
        }
    else:
        report["cas"] = {"column": None, "note": "No common CAS column name detected"}
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path)
    parser.add_argument("--cas-column")
    parser.add_argument("--sheet", help="XLSX worksheet to inspect (defaults to the first)")
    args = parser.parse_args()
    try:
        print(
            json.dumps(
                inspect(args.input, args.cas_column, args.sheet),
                ensure_ascii=False,
                indent=2,
            )
        )
    except (FileNotFoundError, ValueError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
