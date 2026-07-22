"""Safe tabular input and output helpers."""

from __future__ import annotations

import os
import posixpath
import re
import tempfile
import warnings
from copy import copy, deepcopy
from dataclasses import dataclass, field
from io import BytesIO
from numbers import Integral
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

SUPPORTED_INPUTS = {".csv", ".tsv", ".xlsx"}
FORMULA_PREFIXES = ("=", "+", "-", "@")
MAX_EXCEL_COLUMNS = 16_384
_CONTEXT_ATTRIBUTE = "_aromanexus_table_context"
_UNSUPPORTED_PART_MARKERS = {
    "activex": "ActiveX controls",
    "comments": "cell comments",
    "ctrlprops": "form controls",
    "embeddings": "embedded OLE objects",
    "persons": "threaded-comment authors",
    "slicercaches": "slicer caches",
    "slicers": "slicers",
    "threadedcomments": "threaded comments",
    "vml": "VML drawings",
    "_xmlsignatures": "digital signatures",
}
_UNSUPPORTED_DRAWING_ELEMENTS = {"contentPart", "cxnSp", "grpSp", "sp"}
_SAFE_REMOVED_PARTS = {"xl/calcChain.xml", "xl/sharedStrings.xml"}
_SPREADSHEET_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_PACKAGE_RELATIONSHIP_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
_CELL_BLOCK = re.compile(
    rb'<c\b(?=[^>]*\br="(?P<coordinate>[A-Z]{1,3}[1-9][0-9]*)")[^>]*>.*?</c>',
    re.DOTALL,
)
_FORMULA_ELEMENT = re.compile(rb"<f(?:\s[^>]*)?(?:/>|>.*?</f>)", re.DOTALL)
_VALUE_ELEMENT = re.compile(rb"<v(?:\s[^>]*)?(?:/>|>.*?</v>)", re.DOTALL)
_CELL_TYPE_ATTRIBUTE = re.compile(rb'\s+t="(?P<cell_type>[^"]*)"')
_CELL_COORDINATE = re.compile(r"(?P<column>[A-Z]{1,3})(?P<row>[1-9][0-9]*)\Z")


@dataclass(slots=True)
class TableContext:
    """Immutable input snapshot plus the cells intentionally changed by a workflow."""

    source_path: Path
    sheet_name: str | None
    original_columns: tuple[Any, ...]
    row_count: int
    template_bytes: bytes | None = None
    touched_cells: set[tuple[int, Any]] = field(default_factory=set)
    preservation_validated: bool = False
    layout_validated_new_columns: int = -1
    planned_checkpoint_path: Path | None = None
    owned_checkpoint_path: Path | None = None
    owned_checkpoint_identity: tuple[int, int] | None = None
    checkpoint_replace_existing: bool = False

    def record_touch(self, frame: pd.DataFrame, index: Any, column: Any) -> None:
        location = frame.index.get_loc(index)
        if not isinstance(location, Integral):
            raise ValueError("AromaNexus requires uniquely indexed input rows.")
        self.touched_cells.add((int(location), column))


def _validate_source(path: str | Path) -> tuple[Path, str]:
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(f"Input file does not exist: {source}")
    suffix = source.suffix.lower()
    if suffix not in SUPPORTED_INPUTS:
        supported = ", ".join(sorted(SUPPORTED_INPUTS))
        raise ValueError(f"Unsupported input type {suffix!r}; expected one of: {supported}")
    return source, suffix


def _worksheet_names(template: bytes) -> list[str]:
    workbook = load_workbook(
        BytesIO(template),
        read_only=True,
        data_only=False,
        keep_links=True,
        rich_text=True,
    )
    try:
        return [worksheet.title for worksheet in workbook.worksheets]
    finally:
        workbook.close()


def workbook_sheet_names(path: str | Path) -> list[str]:
    """Return XLSX worksheet names in workbook order without modifying the file."""

    source, suffix = _validate_source(path)
    if suffix != ".xlsx":
        return []
    return _worksheet_names(source.read_bytes())


def read_table_with_context(
    path: str | Path,
    *,
    sheet_name: str | None = None,
) -> tuple[pd.DataFrame, TableContext]:
    """Read a table and retain an immutable snapshot for workbook-aware output."""

    source, suffix = _validate_source(path)
    template: bytes | None = None
    selected_sheet: str | None = None
    if suffix == ".xlsx":
        template = source.read_bytes()
        available = _worksheet_names(template)
        if not available:
            raise ValueError(f"XLSX workbook contains no worksheets: {source}")
        selected_sheet = sheet_name or available[0]
        if selected_sheet not in available:
            choices = ", ".join(available)
            raise ValueError(
                f"Worksheet {selected_sheet!r} was not found in {source}. Available: {choices}"
            )
        frame = pd.read_excel(BytesIO(template), sheet_name=selected_sheet, dtype=object)
    else:
        if sheet_name is not None:
            raise ValueError("--sheet is only valid for XLSX input files.")
        separator = "\t" if suffix == ".tsv" else ","
        frame = pd.read_csv(source, sep=separator, dtype=object, keep_default_na=False)

    context = TableContext(
        source_path=source,
        sheet_name=selected_sheet,
        original_columns=tuple(frame.columns),
        row_count=len(frame),
        template_bytes=template,
    )
    frame.attrs[_CONTEXT_ATTRIBUTE] = context
    return frame, context


def read_table(path: str | Path, *, sheet_name: str | None = None) -> pd.DataFrame:
    """Read a supported workbook sheet or delimited text file."""

    frame, _ = read_table_with_context(path, sheet_name=sheet_name)
    frame.attrs.pop(_CONTEXT_ATTRIBUTE, None)
    return frame


def record_touched_cell(frame: pd.DataFrame, index: Any, column: Any) -> None:
    """Record a cell that a workflow intentionally created or replaced."""

    context = frame.attrs.get(_CONTEXT_ATTRIBUTE)
    if isinstance(context, TableContext):
        context.record_touch(frame, index, column)


def require_columns(frame: pd.DataFrame, *columns: str) -> None:
    """Raise once with every missing column name."""

    missing = [column for column in columns if column not in frame.columns]
    if missing:
        available = ", ".join(map(str, frame.columns))
        message = f"Missing required column(s): {', '.join(missing)}. Available: {available}"
        raise ValueError(message)


def derive_output_path(input_path: str | Path, suffix: str) -> Path:
    """Create a sibling output path without touching the source file."""

    source = Path(input_path)
    return source.with_name(f"{source.stem}{suffix}{source.suffix}")


def sanitize_excel_cell(value: Any) -> Any:
    """Prevent untrusted text from becoming an Excel formula."""

    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return f"'{value}"
    return value


def sanitize_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Return a copy whose string cells are safe to open in spreadsheet software."""

    safe = frame.copy()
    for column in safe.columns:
        safe[column] = safe[column].map(sanitize_excel_cell)
    return safe


def _same_path(first: Path, second: Path) -> bool:
    if second.exists():
        try:
            if os.path.samefile(first, second):
                return True
        except OSError:
            pass
    first_path = os.path.normcase(str(first.resolve()))
    second_path = os.path.normcase(str(second.resolve()))
    return first_path == second_path


def _unsupported_parts(template: bytes) -> list[str]:
    features: set[str] = set()
    try:
        with ZipFile(BytesIO(template)) as archive:
            names = archive.namelist()
            for name in names:
                lowered = name.lower()
                for marker, label in _UNSUPPORTED_PART_MARKERS.items():
                    if marker in lowered:
                        features.add(label)
            for name in names:
                lowered = name.lower()
                if not (lowered.startswith("xl/drawings/") and lowered.endswith(".xml")):
                    continue
                root = ElementTree.fromstring(archive.read(name))
                for element in root.iter():
                    local_name = element.tag.rsplit("}", 1)[-1]
                    if local_name in _UNSUPPORTED_DRAWING_ELEMENTS:
                        features.add("drawing shapes")
                        break
    except (BadZipFile, ElementTree.ParseError) as exc:
        raise ValueError(f"Invalid XLSX package: {exc}") from exc
    return sorted(features)


def _openpyxl_roundtrip_issues(template: bytes) -> list[str]:
    workbook = None
    roundtrip = BytesIO()
    try:
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            workbook = load_workbook(
                BytesIO(template),
                read_only=False,
                data_only=False,
                keep_links=True,
                rich_text=True,
            )
            workbook.save(roundtrip)
    finally:
        if workbook is not None:
            workbook.close()

    issues = {
        str(item.message)
        for item in caught
        if "not supported" in str(item.message).lower()
        or "will be removed" in str(item.message).lower()
    }
    with ZipFile(BytesIO(template)) as source_archive:
        source_parts = {name for name in source_archive.namelist() if not name.endswith("/")}
    with ZipFile(BytesIO(roundtrip.getvalue())) as saved_archive:
        saved_parts = {name for name in saved_archive.namelist() if not name.endswith("/")}
    removed_parts = sorted(source_parts - saved_parts - _SAFE_REMOVED_PARTS)
    if removed_parts:
        preview = ", ".join(removed_parts[:8])
        if len(removed_parts) > 8:
            preview += f", and {len(removed_parts) - 8} more"
        issues.add(f"openpyxl round-trip would remove package parts: {preview}")
    return sorted(issues)


def _validate_target_sheet_layout(
    context: TableContext,
    *,
    minimum_new_columns: int,
) -> None:
    if context.template_bytes is None or context.sheet_name is None:
        return
    workbook = load_workbook(
        BytesIO(context.template_bytes),
        read_only=False,
        data_only=False,
        keep_links=True,
        rich_text=True,
    )
    try:
        worksheet = workbook[context.sheet_name]
        if minimum_new_columns and worksheet.max_column + minimum_new_columns > MAX_EXCEL_COLUMNS:
            raise ValueError(
                f"Worksheet {context.sheet_name!r} has no room for AromaNexus output columns "
                f"within Excel's {MAX_EXCEL_COLUMNS:,}-column limit."
            )
        last_table_row = context.row_count + 1
        last_table_column = len(context.original_columns)
        for merged_range in worksheet.merged_cells.ranges:
            intersects_rows = merged_range.min_row <= last_table_row and merged_range.max_row >= 1
            intersects_columns = (
                merged_range.min_col <= last_table_column and merged_range.max_col >= 1
            )
            if intersects_rows and intersects_columns:
                raise ValueError(
                    f"Worksheet {context.sheet_name!r} has merged range {merged_range} inside "
                    "the selected tabular data; unmerge it before enrichment."
                )
    finally:
        workbook.close()


def validate_table_output(
    context: TableContext,
    path: str | Path,
    *,
    force: bool = False,
    minimum_new_columns: int = 0,
) -> Path:
    """Validate an output path and XLSX preservation limits before provider calls."""

    destination = Path(path)
    suffix = destination.suffix.lower()
    if suffix not in SUPPORTED_INPUTS:
        supported = ", ".join(sorted(SUPPORTED_INPUTS))
        raise ValueError(f"Unsupported output type {suffix!r}; expected one of: {supported}")
    if _same_path(context.source_path, destination):
        raise ValueError(
            "Input and output paths must differ; --force cannot replace the source file."
        )
    if destination.exists() and not force:
        raise FileExistsError(f"Output already exists: {destination}. Pass --force to replace it.")
    preserve_xlsx = (
        suffix == ".xlsx"
        and context.source_path.suffix.lower() == ".xlsx"
        and context.template_bytes is not None
    )
    if preserve_xlsx and minimum_new_columns > context.layout_validated_new_columns:
        _validate_target_sheet_layout(
            context,
            minimum_new_columns=minimum_new_columns,
        )
        context.layout_validated_new_columns = minimum_new_columns
    if preserve_xlsx and not context.preservation_validated:
        unsupported = _unsupported_parts(context.template_bytes)
        roundtrip_issues = _openpyxl_roundtrip_issues(context.template_bytes)
        if unsupported or roundtrip_issues:
            details = unsupported + roundtrip_issues
            raise ValueError(
                "XLSX contains features that cannot be preserved safely: " + "; ".join(details)
            )
        context.preservation_validated = True
    return destination


def _worksheet_part_map(archive: ZipFile) -> dict[str, str]:
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    relationships_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {
        relationship.attrib["Id"]: relationship.attrib["Target"]
        for relationship in relationships_root.findall(
            f"{{{_PACKAGE_RELATIONSHIP_NAMESPACE}}}Relationship"
        )
    }
    parts: dict[str, str] = {}
    for sheet in workbook_root.findall(
        f"{{{_SPREADSHEET_NAMESPACE}}}sheets/{{{_SPREADSHEET_NAMESPACE}}}sheet"
    ):
        relationship_id = sheet.attrib[f"{{{_OFFICE_RELATIONSHIP_NAMESPACE}}}id"]
        target = targets[relationship_id].replace("\\", "/")
        if target.startswith("/"):
            part = target.lstrip("/")
        else:
            part = posixpath.normpath(posixpath.join("xl", target))
        parts[sheet.attrib["name"]] = part
    return parts


def _formula_cached_values(worksheet_xml: bytes) -> dict[bytes, tuple[bytes, bytes | None]]:
    cached: dict[bytes, tuple[bytes, bytes | None]] = {}
    root = ElementTree.fromstring(worksheet_xml)
    for cell in root.iter():
        if cell.tag.rsplit("}", 1)[-1] != "c":
            continue
        coordinate = cell.attrib.get("r")
        formula = None
        value = None
        for child in cell:
            local_name = child.tag.rsplit("}", 1)[-1]
            if local_name == "f":
                formula = child
            elif local_name == "v":
                value = child
        if coordinate and formula is not None and value is not None:
            cached_value = ElementTree.Element("v")
            cached_value.text = value.text
            cached[coordinate.encode("ascii")] = (
                ElementTree.tostring(
                    cached_value,
                    encoding="utf-8",
                    short_empty_elements=False,
                )
                if value.text is not None
                else b"<v></v>",
                cell.attrib.get("t", "").encode("utf-8") or None,
            )
    return cached


def _restore_cached_values(
    worksheet_xml: bytes,
    cached_values: dict[bytes, tuple[bytes, bytes | None]],
    touched_coordinates: set[str],
) -> bytes:
    touched = {coordinate.encode("ascii") for coordinate in touched_coordinates}

    def restore(match: re.Match[bytes]) -> bytes:
        coordinate = match.group("coordinate")
        block = match.group(0)
        cached = cached_values.get(coordinate)
        if coordinate in touched or cached is None or _FORMULA_ELEMENT.search(block) is None:
            return block
        cached_value, cached_type = cached
        opening_end = block.index(b">")
        opening_tag = _CELL_TYPE_ATTRIBUTE.sub(b"", block[:opening_end])
        if cached_type is not None:
            opening_tag += b' t="' + cached_type + b'"'
        block = opening_tag + block[opening_end:]
        current = _VALUE_ELEMENT.search(block)
        if current is not None:
            return block[: current.start()] + cached_value + block[current.end() :]
        return block[:-4] + cached_value + block[-4:]

    return _CELL_BLOCK.sub(restore, worksheet_xml)


def _worksheet_tag(local_name: str) -> str:
    return f"{{{_SPREADSHEET_NAMESPACE}}}{local_name}"


def _cell_position(coordinate: str) -> tuple[int, int]:
    match = _CELL_COORDINATE.fullmatch(coordinate)
    if match is None:
        raise ValueError(f"Invalid XLSX cell coordinate: {coordinate!r}")
    column = 0
    for character in match.group("column"):
        column = column * 26 + ord(character) - ord("A") + 1
    return int(match.group("row")), column


def _row_cells(row: ElementTree.Element) -> dict[str, ElementTree.Element]:
    cells: dict[str, ElementTree.Element] = {}
    for cell in row.findall(_worksheet_tag("c")):
        coordinate = cell.attrib.get("r")
        if coordinate is None:
            raise ValueError("XLSX worksheet contains a cell without a coordinate.")
        cell_row, _ = _cell_position(coordinate)
        row_number = row.attrib.get("r")
        if row_number is None or cell_row != int(row_number):
            raise ValueError(f"XLSX cell {coordinate!r} is stored in the wrong row.")
        if coordinate in cells:
            raise ValueError(f"XLSX worksheet contains duplicate cell {coordinate!r}.")
        cells[coordinate] = cell
    return cells


def _insert_cell(row: ElementTree.Element, cell: ElementTree.Element) -> None:
    _, target_column = _cell_position(cell.attrib["r"])
    for position, child in enumerate(row):
        if child.tag != _worksheet_tag("c"):
            row.insert(position, cell)
            return
        _, child_column = _cell_position(child.attrib["r"])
        if child_column > target_column:
            row.insert(position, cell)
            return
    row.append(cell)


def _copy_style_reference(
    source_cell: ElementTree.Element,
    saved_cell: ElementTree.Element,
) -> None:
    style_id = source_cell.attrib.get("s")
    if style_id is None:
        saved_cell.attrib.pop("s", None)
    else:
        saved_cell.attrib["s"] = style_id


def _blank_cell_with_source_style(source_cell: ElementTree.Element) -> ElementTree.Element | None:
    style_id = source_cell.attrib.get("s")
    if style_id is None:
        return None
    return ElementTree.Element(
        _worksheet_tag("c"),
        {"r": source_cell.attrib["r"], "s": style_id},
    )


def _shared_formula_cells_affected_by_touches(
    source_data: ElementTree.Element,
    touched_coordinates: set[str],
) -> set[str]:
    groups: dict[str, set[str]] = {}
    for row in source_data.findall(_worksheet_tag("row")):
        for coordinate, cell in _row_cells(row).items():
            formula = cell.find(_worksheet_tag("f"))
            if formula is None or formula.attrib.get("t") != "shared":
                continue
            shared_index = formula.attrib.get("si")
            if shared_index is None:
                raise ValueError(f"Shared formula cell {coordinate!r} has no shared index.")
            groups.setdefault(shared_index, set()).add(coordinate)
    return {
        coordinate
        for coordinates in groups.values()
        if coordinates & touched_coordinates
        for coordinate in coordinates
    }


def _expand_worksheet_dimension(root: ElementTree.Element) -> None:
    dimension = root.find(_worksheet_tag("dimension"))
    sheet_data = root.find(_worksheet_tag("sheetData"))
    if dimension is None or sheet_data is None:
        return
    coordinates = [
        _cell_position(cell.attrib["r"])
        for row in sheet_data.findall(_worksheet_tag("row"))
        for cell in row.findall(_worksheet_tag("c"))
    ]
    if not coordinates:
        return
    rows, columns = zip(*coordinates, strict=True)
    reference = dimension.attrib.get("ref", "A1")
    endpoints = reference.split(":", 1)
    dimension_coordinates = [_cell_position(item.replace("$", "")) for item in endpoints]
    dimension_rows, dimension_columns = zip(*dimension_coordinates, strict=True)
    minimum_row = min(*rows, *dimension_rows)
    maximum_row = max(*rows, *dimension_rows)
    minimum_column = min(*columns, *dimension_columns)
    maximum_column = max(*columns, *dimension_columns)
    first = f"{get_column_letter(minimum_column)}{minimum_row}"
    last = f"{get_column_letter(maximum_column)}{maximum_row}"
    dimension.attrib["ref"] = first if first == last else f"{first}:{last}"


def _merge_original_cells(
    source_xml: bytes,
    saved_xml: bytes,
    touched_coordinates: set[str],
) -> bytes:
    """Restore source cells while retaining intentional values from the saved sheet."""

    source_root = ElementTree.fromstring(source_xml)
    saved_root = ElementTree.fromstring(saved_xml)
    source_data = source_root.find(_worksheet_tag("sheetData"))
    saved_data = saved_root.find(_worksheet_tag("sheetData"))
    if source_data is None or saved_data is None:
        raise ValueError("XLSX worksheet is missing sheetData.")

    saved_rows: dict[int, ElementTree.Element] = {}
    for row in saved_data.findall(_worksheet_tag("row")):
        row_number = row.attrib.get("r")
        if row_number is None:
            raise ValueError("XLSX worksheet contains a row without a number.")
        numeric_row = int(row_number)
        if numeric_row in saved_rows:
            raise ValueError(f"XLSX worksheet contains duplicate row {numeric_row}.")
        saved_rows[numeric_row] = row

    touched = set(touched_coordinates)
    keep_saved_formula_cells = _shared_formula_cells_affected_by_touches(source_data, touched)
    for source_row in source_data.findall(_worksheet_tag("row")):
        row_number = source_row.attrib.get("r")
        if row_number is None:
            raise ValueError("XLSX worksheet contains a row without a number.")
        numeric_row = int(row_number)
        source_cells = _row_cells(source_row)
        saved_row = saved_rows.get(numeric_row)

        if saved_row is None:
            missing_saved_formulas = (set(source_cells) & keep_saved_formula_cells) - touched
            if missing_saved_formulas:
                coordinate = min(missing_saved_formulas, key=_cell_position)
                raise ValueError(
                    f"Shared formula cell {coordinate!r} was lost during XLSX preservation."
                )
            restored_row = deepcopy(source_row)
            restored_cells = _row_cells(restored_row)
            for coordinate, restored_cell in list(restored_cells.items()):
                if coordinate not in touched:
                    continue
                position = list(restored_row).index(restored_cell)
                restored_row.remove(restored_cell)
                blank = _blank_cell_with_source_style(source_cells[coordinate])
                if blank is not None:
                    restored_row.insert(position, blank)
            for position, candidate in enumerate(saved_data.findall(_worksheet_tag("row"))):
                if int(candidate.attrib["r"]) > numeric_row:
                    saved_data.insert(position, restored_row)
                    break
            else:
                saved_data.append(restored_row)
            saved_rows[numeric_row] = restored_row
            continue

        saved_cells = _row_cells(saved_row)
        for coordinate, source_cell in source_cells.items():
            saved_cell = saved_cells.get(coordinate)
            if coordinate in touched or coordinate in keep_saved_formula_cells:
                if saved_cell is not None:
                    _copy_style_reference(source_cell, saved_cell)
                elif coordinate in keep_saved_formula_cells and coordinate not in touched:
                    raise ValueError(
                        f"Shared formula cell {coordinate!r} was lost during XLSX preservation."
                    )
                else:
                    blank = _blank_cell_with_source_style(source_cell)
                    if blank is not None:
                        _insert_cell(saved_row, blank)
                continue

            restored_cell = deepcopy(source_cell)
            if saved_cell is None:
                _insert_cell(saved_row, restored_cell)
                continue
            position = list(saved_row).index(saved_cell)
            saved_row.remove(saved_cell)
            saved_row.insert(position, restored_cell)

    _expand_worksheet_dimension(saved_root)
    ElementTree.register_namespace("", _SPREADSHEET_NAMESPACE)
    ElementTree.register_namespace("r", _OFFICE_RELATIONSHIP_NAMESPACE)
    return ElementTree.tostring(saved_root, encoding="utf-8", xml_declaration=True)


def _sheet_relationship_part(worksheet_part: str) -> str:
    directory, filename = posixpath.split(worksheet_part)
    return posixpath.join(directory, "_rels", f"{filename}.rels")


def _remove_calc_chain_reference(part_name: str, payload: bytes) -> bytes:
    if part_name == "[Content_Types].xml":
        root = ElementTree.fromstring(payload)
        changed = False
        for child in list(root):
            part = child.attrib.get("PartName", "")
            content_type = child.attrib.get("ContentType", "")
            if part.casefold() == "/xl/calcchain.xml" or content_type.endswith("calcChain+xml"):
                root.remove(child)
                changed = True
        if changed:
            return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
    elif part_name == "xl/_rels/workbook.xml.rels":
        root = ElementTree.fromstring(payload)
        changed = False
        for child in list(root):
            if child.attrib.get("Type", "").endswith("/calcChain"):
                root.remove(child)
                changed = True
        if changed:
            return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
    return payload


def _build_preserved_package(
    template: bytes,
    saved: bytes,
    *,
    selected_sheet: str,
    touched_coordinates: set[str],
) -> bytes:
    replacements: dict[str, bytes] = {}
    with ZipFile(BytesIO(template)) as source_archive, ZipFile(BytesIO(saved)) as saved_archive:
        source_parts = _worksheet_part_map(source_archive)
        saved_parts = _worksheet_part_map(saved_archive)
        source_part = source_parts[selected_sheet]
        saved_part = saved_parts[selected_sheet]
        if posixpath.dirname(source_part) != posixpath.dirname(saved_part):
            raise ValueError("XLSX worksheet relationship paths changed during preservation.")
        source_worksheet = source_archive.read(source_part)
        cached = _formula_cached_values(source_worksheet)
        saved_worksheet = _restore_cached_values(
            saved_archive.read(saved_part),
            cached,
            touched_coordinates,
        )
        replacements[source_part] = _merge_original_cells(
            source_worksheet,
            saved_worksheet,
            touched_coordinates,
        )
        source_relationships = _sheet_relationship_part(source_part)
        saved_relationships = _sheet_relationship_part(saved_part)
        if saved_relationships in saved_archive.namelist():
            replacements[source_relationships] = saved_archive.read(saved_relationships)
        elif source_relationships in source_archive.namelist():
            raise ValueError("XLSX worksheet relationships were lost during preservation.")

        rewritten = BytesIO()
        with ZipFile(rewritten, "w") as output_archive:
            written: set[str] = set()
            for item in source_archive.infolist():
                if item.filename == "xl/calcChain.xml":
                    continue
                payload = replacements.get(item.filename, source_archive.read(item.filename))
                payload = _remove_calc_chain_reference(item.filename, payload)
                output_archive.writestr(
                    item,
                    payload,
                )
                written.add(item.filename)
            for part_name, payload in replacements.items():
                if part_name not in written:
                    output_archive.writestr(part_name, payload)
    return rewritten.getvalue()


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if bool(pd.isna(value)):
            return None
    except (TypeError, ValueError):
        pass
    return sanitize_excel_cell(value)


def _column_width(frame: pd.DataFrame, column: Any) -> float:
    values = [column, *frame[column].tolist()]
    longest = max((len(str(value)) for value in values if value is not None), default=0)
    return float(min(max(longest + 2, 12), 50))


def _write_preserved_xlsx(
    frame: pd.DataFrame,
    temporary: Path,
    context: TableContext,
) -> None:
    if context.template_bytes is None or context.sheet_name is None:
        raise ValueError("XLSX preservation context is incomplete.")
    original_count = len(context.original_columns)
    if tuple(frame.columns[:original_count]) != context.original_columns:
        raise ValueError(
            "A workflow reordered or removed original columns; refusing unsafe XLSX output."
        )
    if len(frame.columns) > MAX_EXCEL_COLUMNS:
        raise ValueError(f"XLSX output exceeds Excel's {MAX_EXCEL_COLUMNS:,}-column limit.")

    workbook = load_workbook(
        BytesIO(context.template_bytes),
        read_only=False,
        data_only=False,
        keep_links=True,
        rich_text=True,
    )
    try:
        worksheet = workbook[context.sheet_name]
        existing_positions = {
            column: position for position, column in enumerate(context.original_columns, 1)
        }
        new_columns = list(frame.columns[original_count:])
        append_start = max(original_count, worksheet.max_column) + 1
        if append_start - 1 + len(new_columns) > MAX_EXCEL_COLUMNS:
            raise ValueError(f"XLSX output exceeds Excel's {MAX_EXCEL_COLUMNS:,}-column limit.")
        new_positions = {column: append_start + offset for offset, column in enumerate(new_columns)}

        for column, position in new_positions.items():
            header = worksheet.cell(row=1, column=position)
            header.value = _excel_value(column)
            if original_count:
                source_header = worksheet.cell(row=1, column=original_count)
                if source_header.has_style:
                    header._style = copy(source_header._style)
            column_dimension = worksheet.column_dimensions[get_column_letter(position)]
            # Excel-generated files can use a non-zero base style. A fresh
            # ColumnDimension otherwise registers a new all-zero style whose
            # index exists only in openpyxl's rewritten styles.xml. Reuse the
            # workbook's real base style so the immutable source styles part
            # remains valid and byte-preservable.
            column_dimension._style = copy(workbook._cell_styles[0])
            column_dimension.width = _column_width(frame, column)

        frame_positions = {column: position for position, column in enumerate(frame.columns)}
        # ``DataFrame.iat`` extracts a Series for every scalar access. Pandas then
        # deep-copies ``frame.attrs``, including our growing XLSX context. A single
        # object view avoids that quadratic checkpoint tax without copying values.
        frame_values = frame.to_numpy(dtype=object, copy=False)
        touched_coordinates: set[str] = set()
        for row_position, column in sorted(
            context.touched_cells,
            key=lambda item: (item[0], str(item[1])),
        ):
            output_column = existing_positions.get(column, new_positions.get(column))
            frame_column = frame_positions.get(column)
            if output_column is None or frame_column is None:
                raise ValueError(f"Touched column {column!r} is missing from the output frame.")
            cell = worksheet.cell(row=row_position + 2, column=output_column)
            if isinstance(cell, MergedCell):
                raise ValueError(
                    f"Cannot write {column!r} at row {row_position + 2}: "
                    "cell is inside a merged range."
                )
            cell.value = _excel_value(frame_values[row_position, frame_column])
            touched_coordinates.add(f"{get_column_letter(output_column)}{row_position + 2}")

        workbook.save(temporary)
    finally:
        workbook.close()

    temporary.write_bytes(
        _build_preserved_package(
            context.template_bytes,
            temporary.read_bytes(),
            selected_sheet=context.sheet_name,
            touched_coordinates=touched_coordinates,
        )
    )

    validation = load_workbook(
        temporary,
        read_only=True,
        data_only=False,
        keep_links=True,
        rich_text=True,
    )
    try:
        if context.sheet_name not in validation.sheetnames:
            raise ValueError(f"Saved XLSX is missing worksheet {context.sheet_name!r}.")
    finally:
        validation.close()


def _commit_temporary(
    temporary: Path,
    destination: Path,
    *,
    force: bool,
) -> None:
    if force:
        os.replace(temporary, destination)
        return
    try:
        if os.name == "nt":
            os.rename(temporary, destination)
        else:
            os.link(temporary, destination)
            temporary.unlink()
    except FileExistsError as exc:
        raise FileExistsError(
            f"Output already exists: {destination}. Pass --force to replace it."
        ) from exc


def write_table(
    frame: pd.DataFrame,
    path: str | Path,
    *,
    force: bool = False,
    context: TableContext | None = None,
) -> Path:
    """Atomically write a table, preserving a source XLSX when context is supplied."""

    destination = Path(path)
    suffix = destination.suffix.lower()
    if suffix not in SUPPORTED_INPUTS:
        supported = ", ".join(sorted(SUPPORTED_INPUTS))
        raise ValueError(f"Unsupported output type {suffix!r}; expected one of: {supported}")
    if context is not None:
        validate_table_output(context, destination, force=force)
    elif destination.exists() and not force:
        raise FileExistsError(f"Output already exists: {destination}. Pass --force to replace it.")
    destination.parent.mkdir(parents=True, exist_ok=True)
    handle = tempfile.NamedTemporaryFile(
        prefix=f".{destination.stem}-",
        suffix=destination.suffix,
        dir=destination.parent,
        delete=False,
    )
    temporary = Path(handle.name)
    handle.close()
    try:
        preserve_xlsx = (
            suffix == ".xlsx"
            and context is not None
            and context.source_path.suffix.lower() == ".xlsx"
            and context.template_bytes is not None
        )
        if preserve_xlsx:
            _write_preserved_xlsx(frame, temporary, context)
        else:
            safe = sanitize_frame(frame)
            if suffix == ".xlsx":
                safe.to_excel(temporary, index=False)
            else:
                separator = "\t" if suffix == ".tsv" else ","
                safe.to_csv(temporary, sep=separator, index=False, encoding="utf-8-sig")
        _commit_temporary(temporary, destination, force=force)
    finally:
        temporary.unlink(missing_ok=True)
    return destination
