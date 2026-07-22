import os
import re
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import openpyxl
import pandas as pd
import pytest
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from aromanexus import excel_io
from aromanexus.models import LookupResult
from aromanexus.workflows import PUBCHEM_COLUMN_MAP, PUBCHEM_ODOR_KEYS, run_pubchem


class NoLookupPubChem:
    def __init__(self) -> None:
        self.calls = 0

    def lookup(self, identifier, include_odor=True):
        self.calls += 1
        raise AssertionError(f"Unexpected PubChem lookup: {identifier!r}")


class FormulaPubChem:
    def lookup(self, identifier, include_odor=True):
        return LookupResult(
            provider="PubChem",
            values={
                "title": '=HYPERLINK("https://example.test")',
                "cas_numbers": ["110-54-3"],
            },
        )


class InterruptedPubChem:
    def __init__(self) -> None:
        self.calls = 0

    def lookup(self, identifier, include_odor=True):
        self.calls += 1
        if self.calls == 2:
            raise RuntimeError("simulated interruption")
        return LookupResult(
            provider="PubChem",
            values={"title": "First result", "cas_numbers": ["110-54-3"]},
        )


def _silent(*_):
    return None


def _create_preservation_workbook(path: Path, *, cover_first: bool = False) -> None:
    workbook = openpyxl.Workbook()
    data = workbook.active
    data.title = "Data"
    if cover_first:
        cover = workbook.create_sheet("Cover", 0)
        cover["A1"] = "AromaNexus preservation fixture"
        cover["A2"] = "Do not modify"

    data.append(["Name", "Measured value", "Calculated value"])
    data.append(["C6", 10, "=B2*2"])
    data.append(["C7", 20, "=B3*2"])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    row_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="000000")
    for cell in data[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)
    for cell in data[2]:
        cell.fill = row_fill
    data["B2"].number_format = "0.00"
    data.column_dimensions["A"].width = 18
    data.column_dimensions["A"].font = Font(italic=True)
    data.column_dimensions["B"].width = 22
    data.column_dimensions["C"].width = 24
    data.row_dimensions[1].height = 28
    data.row_dimensions[2].height = 21
    data.freeze_panes = "A2"
    data.auto_filter.ref = "A1:C3"

    table = Table(displayName="DataTable", ref="A1:C3")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    data.add_table(table)

    validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="100",
    )
    validation.add("B2:B3")
    data.add_data_validation(validation)
    data.conditional_formatting.add(
        "B2:B3",
        CellIsRule(
            operator="greaterThan",
            formula=["15"],
            fill=PatternFill("solid", fgColor="C6EFCE"),
        ),
    )

    notes = workbook.create_sheet("Notes")
    notes.append(["Field", "Value"])
    notes.append(["Owner", "Tianyuan Chen"])
    notes.append(["Name length", "=LEN(B2)"])
    notes.merge_cells("A4:B4")
    notes["A4"] = "Preservation regression fixture"

    workbook.properties.title = "AromaNexus preservation fixture"
    workbook.properties.creator = "Tianyuan Chen"
    workbook.save(path)
    workbook.close()


def _header_positions(worksheet) -> dict[object, int]:
    return {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}


def _rewrite_xlsx(
    path: Path,
    replacements: dict[str, bytes],
    additions: dict[str, bytes] | None = None,
) -> None:
    rewritten = BytesIO()
    with ZipFile(BytesIO(path.read_bytes())) as source, ZipFile(rewritten, "w") as output:
        existing = set(source.namelist())
        for item in source.infolist():
            output.writestr(item, replacements.get(item.filename, source.read(item.filename)))
        for name, payload in (additions or {}).items():
            if name not in existing:
                output.writestr(name, payload)
    path.write_bytes(rewritten.getvalue())


def _xlsx_part(path: Path, name: str) -> bytes:
    with ZipFile(path) as archive:
        return archive.read(name)


def _inject_formula_cache(
    worksheet_xml: bytes,
    coordinate: str,
    value: str,
    *,
    cell_type: str | None = None,
) -> bytes:
    pattern = re.compile(
        rb'<c\b(?=[^>]*\br="' + coordinate.encode("ascii") + rb'")[^>]*>.*?</c>',
        re.DOTALL,
    )

    def replace(match: re.Match[bytes]) -> bytes:
        block = match.group(0)
        assert b"<f" in block
        opening_end = block.index(b">")
        opening = re.sub(rb'\s+t="[^"]*"', b"", block[:opening_end])
        if cell_type is not None:
            opening += f' t="{cell_type}"'.encode()
        block = opening + block[opening_end:]
        cached = f"<v>{value}</v>".encode()
        block, count = re.subn(rb"<v(?:\s[^>]*)?(?:/>|>.*?</v>)", cached, block, count=1)
        assert count == 1
        return block

    result, count = pattern.subn(replace, worksheet_xml, count=1)
    assert count == 1
    return result


def test_all_skipped_xlsx_run_preserves_workbook_structure(tmp_path: Path):
    source = tmp_path / "preservation_input.xlsx"
    destination = tmp_path / "preservation_output.xlsx"
    _create_preservation_workbook(source)
    client = NoLookupPubChem()

    run_pubchem(
        source,
        client,
        output_path=destination,
        sheet_name="Data",
        identifier_column="Name",
        skip_patterns=[r"^C\d+$"],
        include_odor=False,
        include_provenance=False,
        checkpoint_every=1,
        progress=_silent,
    )

    assert client.calls == 0
    assert not (tmp_path / "preservation_output.partial.xlsx").exists()
    workbook = openpyxl.load_workbook(destination, data_only=False)
    try:
        assert workbook.sheetnames == ["Data", "Notes"]
        assert workbook.properties.title == "AromaNexus preservation fixture"
        assert workbook.properties.creator == "Tianyuan Chen"
        data = workbook["Data"]
        notes = workbook["Notes"]
        assert [data.cell(1, column).value for column in range(1, 4)] == [
            "Name",
            "Measured value",
            "Calculated value",
        ]
        assert [data.cell(row, 1).value for row in range(2, 4)] == ["C6", "C7"]
        assert data["C2"].value == "=B2*2"
        assert data["C3"].value == "=B3*2"
        assert data["C2"].data_type == "f"
        assert data["A1"].font.bold is True
        assert data["A1"].fill.fgColor.rgb == "001F4E78"
        assert data["A2"].fill.fgColor.rgb == "00FFF2CC"
        assert data["A1"].border.bottom.style == "thin"
        assert data["B2"].number_format == "0.00"
        assert data.column_dimensions["A"].width == 18
        assert data.column_dimensions["A"].font.italic is True
        assert data.column_dimensions["B"].width == 22
        assert data.column_dimensions["C"].width == 24
        assert data.row_dimensions[1].height == 28
        assert data.row_dimensions[2].height == 21
        assert data.freeze_panes == "A2"
        assert data.auto_filter.ref == "A1:C3"
        assert list(data.tables) == ["DataTable"]
        assert data.tables["DataTable"].ref == "A1:C3"
        validation = data.data_validations.dataValidation[0]
        assert (str(validation.sqref), validation.type, validation.operator) == (
            "B2:B3",
            "whole",
            "between",
        )
        conditional_range = next(iter(data.conditional_formatting))
        assert str(conditional_range.sqref) == "B2:B3"
        assert notes["B2"].value == "Tianyuan Chen"
        assert notes["B3"].value == "=LEN(B2)"
        assert notes["B3"].data_type == "f"
        assert "A4:B4" in {str(item) for item in notes.merged_cells.ranges}
        headers = _header_positions(data)
        assert set(headers) >= {
            "PubChem CAS Resolution",
            "PubChem CAS Candidate Count",
            "Resolved CAS",
        }
    finally:
        workbook.close()


def test_formula_injection_applies_only_to_new_enrichment(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    destination = tmp_path / "output.xlsx"
    _create_preservation_workbook(source)

    run_pubchem(
        source,
        FormulaPubChem(),
        output_path=destination,
        sheet_name="Data",
        identifier_column="Name",
        include_odor=False,
        include_provenance=False,
        checkpoint_every=0,
        progress=_silent,
    )

    workbook = openpyxl.load_workbook(destination, data_only=False)
    try:
        data = workbook["Data"]
        title_column = _header_positions(data)["PubChem Title"]
        assert data["C2"].value == "=B2*2"
        assert data["C2"].data_type == "f"
        assert data.cell(2, title_column).value == '\'=HYPERLINK("https://example.test")'
        assert data.cell(2, title_column).data_type == "s"
    finally:
        workbook.close()


def test_explicit_sheet_selects_non_first_worksheet(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    destination = tmp_path / "output.xlsx"
    _create_preservation_workbook(source, cover_first=True)
    client = NoLookupPubChem()

    run_pubchem(
        source,
        client,
        output_path=destination,
        sheet_name="Data",
        identifier_column="Name",
        skip_patterns=[r"^C\d+$"],
        include_provenance=False,
        checkpoint_every=0,
        progress=_silent,
    )

    workbook = openpyxl.load_workbook(destination, data_only=False)
    try:
        assert workbook.sheetnames == ["Cover", "Data", "Notes"]
        assert workbook["Cover"]["A2"].value == "Do not modify"
        assert workbook["Cover"].max_column == 1
        assert "PubChem CAS Resolution" in _header_positions(workbook["Data"])
    finally:
        workbook.close()


def test_invalid_sheet_and_same_path_fail_before_provider_calls(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    _create_preservation_workbook(source)
    original = source.read_bytes()
    client = NoLookupPubChem()

    with pytest.raises(ValueError, match="Available: Data, Notes"):
        run_pubchem(
            source,
            client,
            output_path=tmp_path / "missing.xlsx",
            sheet_name="Missing",
            progress=_silent,
        )
    with pytest.raises(ValueError, match="Input and output paths must differ"):
        run_pubchem(
            source,
            client,
            output_path=source,
            sheet_name="Data",
            force=True,
            progress=_silent,
        )

    assert client.calls == 0
    assert source.read_bytes() == original


def test_interrupted_checkpoint_is_a_complete_workbook(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    destination = tmp_path / "output.xlsx"
    partial = tmp_path / "output.partial.xlsx"
    _create_preservation_workbook(source)

    with pytest.raises(RuntimeError, match="simulated interruption"):
        run_pubchem(
            source,
            InterruptedPubChem(),
            output_path=destination,
            sheet_name="Data",
            identifier_column="Name",
            include_odor=False,
            include_provenance=False,
            checkpoint_every=1,
            progress=_silent,
        )

    assert not destination.exists()
    assert partial.exists()
    workbook = openpyxl.load_workbook(partial, data_only=False)
    try:
        assert workbook.sheetnames == ["Data", "Notes"]
        data = workbook["Data"]
        assert data["C2"].value == "=B2*2"
        assert data.freeze_panes == "A2"
        assert data.tables["DataTable"].ref == "A1:C3"
        title_column = _header_positions(data)["PubChem Title"]
        assert data.cell(2, title_column).value == "First result"
        assert data.cell(3, title_column).value is None
    finally:
        workbook.close()


def test_unsupported_comments_fail_before_provider_calls(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    destination = tmp_path / "output.xlsx"
    _create_preservation_workbook(source)
    workbook = openpyxl.load_workbook(source)
    workbook["Data"]["A2"].comment = Comment("Keep this formatting", "AromaNexus")
    workbook.save(source)
    workbook.close()
    client = NoLookupPubChem()

    with pytest.raises(ValueError, match="cell comments"):
        run_pubchem(
            source,
            client,
            output_path=destination,
            sheet_name="Data",
            progress=_silent,
        )

    assert client.calls == 0
    assert not destination.exists()


def test_csv_formula_safety_is_unchanged_and_sheet_is_rejected(tmp_path: Path):
    source = tmp_path / "input.csv"
    destination = tmp_path / "output.csv"
    source.write_text("Name,Original\nC6,=1+1\n", encoding="utf-8")
    client = NoLookupPubChem()

    run_pubchem(
        source,
        client,
        output_path=destination,
        identifier_column="Name",
        skip_patterns=[r"^C\d+$"],
        include_provenance=False,
        checkpoint_every=0,
        progress=_silent,
    )

    assert destination.read_bytes().startswith(b"\xef\xbb\xbf")
    output = pd.read_csv(destination, keep_default_na=False)
    assert output.loc[0, "Original"] == "'=1+1"
    assert client.calls == 0

    invalid_destination = tmp_path / "invalid.csv"
    with pytest.raises(ValueError, match="only valid for XLSX"):
        run_pubchem(
            source,
            client,
            output_path=invalid_destination,
            sheet_name="Data",
            progress=_silent,
        )
    assert not invalid_destination.exists()


def test_formula_caches_and_non_target_package_parts_are_preserved(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    destination = tmp_path / "output.xlsx"
    _create_preservation_workbook(source)
    workbook = openpyxl.load_workbook(source)
    workbook["Notes"]["B3"] = '="cached note"'
    workbook.save(source)
    workbook.close()

    data_xml = _inject_formula_cache(
        _xlsx_part(source, "xl/worksheets/sheet1.xml"),
        "C2",
        "20",
    )
    assert b'r="C2"' in data_xml
    data_xml = data_xml.replace(b'r="C2"', b"r='C2'", 1)
    notes_xml = _inject_formula_cache(
        _xlsx_part(source, "xl/worksheets/sheet2.xml"),
        "B3",
        "cached note",
        cell_type="str",
    )
    _rewrite_xlsx(
        source,
        {
            "xl/worksheets/sheet1.xml": data_xml,
            "xl/worksheets/sheet2.xml": notes_xml,
        },
    )
    original_workbook_xml = _xlsx_part(source, "xl/workbook.xml")
    original_styles_xml = _xlsx_part(source, "xl/styles.xml")
    original_notes_xml = _xlsx_part(source, "xl/worksheets/sheet2.xml")

    run_pubchem(
        source,
        NoLookupPubChem(),
        output_path=destination,
        sheet_name="Data",
        identifier_column="Name",
        skip_patterns=[r"^C\d+$"],
        include_odor=False,
        include_provenance=False,
        checkpoint_every=0,
        progress=_silent,
    )

    assert _xlsx_part(destination, "xl/workbook.xml") == original_workbook_xml
    assert _xlsx_part(destination, "xl/styles.xml") == original_styles_xml
    assert _xlsx_part(destination, "xl/worksheets/sheet2.xml") == original_notes_xml
    formulas = openpyxl.load_workbook(destination, data_only=False)
    cached = openpyxl.load_workbook(destination, data_only=True)
    try:
        assert formulas["Data"]["C2"].value == "=B2*2"
        assert cached["Data"]["C2"].value == 20
        assert formulas["Notes"]["B3"].value == '="cached note"'
        assert cached["Notes"]["B3"].value == "cached note"
    finally:
        formulas.close()
        cached.close()


def test_overwritten_formula_does_not_regain_stale_cache(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    destination = tmp_path / "output.xlsx"
    _create_preservation_workbook(source)
    data_xml = _inject_formula_cache(
        _xlsx_part(source, "xl/worksheets/sheet1.xml"),
        "C2",
        "20",
    )
    _rewrite_xlsx(source, {"xl/worksheets/sheet1.xml": data_xml})

    run_pubchem(
        source,
        FormulaPubChem(),
        output_path=destination,
        sheet_name="Data",
        identifier_column="Name",
        resolved_cas_column="Calculated value",
        include_odor=False,
        include_provenance=False,
        checkpoint_every=0,
        progress=_silent,
    )

    formulas = openpyxl.load_workbook(destination, data_only=False)
    cached = openpyxl.load_workbook(destination, data_only=True)
    try:
        assert formulas["Data"]["C2"].value == "110-54-3"
        assert formulas["Data"]["C2"].data_type == "s"
        assert cached["Data"]["C2"].value == "110-54-3"
    finally:
        formulas.close()
        cached.close()


def test_unrelated_partial_files_are_never_deleted(tmp_path: Path):
    source = tmp_path / "result.partial.xlsx"
    destination = tmp_path / "result.xlsx"
    unrelated_partial = tmp_path / "short.partial.xlsx"
    short_destination = tmp_path / "short.xlsx"
    _create_preservation_workbook(source)
    original_source = source.read_bytes()

    run_pubchem(
        source,
        NoLookupPubChem(),
        output_path=destination,
        sheet_name="Data",
        identifier_column="Name",
        skip_patterns=[r"^C\d+$"],
        include_provenance=False,
        checkpoint_every=0,
        progress=_silent,
    )
    assert source.read_bytes() == original_source

    unrelated_partial.write_bytes(b"user-owned checkpoint sentinel")
    run_pubchem(
        source,
        NoLookupPubChem(),
        output_path=short_destination,
        sheet_name="Data",
        identifier_column="Name",
        skip_patterns=[r"^C\d+$"],
        include_provenance=False,
        checkpoint_every=25,
        progress=_silent,
    )
    assert unrelated_partial.read_bytes() == b"user-owned checkpoint sentinel"


def test_existing_enabled_checkpoint_fails_before_provider(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    destination = tmp_path / "output.xlsx"
    partial = tmp_path / "output.partial.xlsx"
    _create_preservation_workbook(source)
    partial.write_bytes(b"user-owned checkpoint sentinel")
    client = NoLookupPubChem()

    with pytest.raises(FileExistsError, match="output.partial.xlsx"):
        run_pubchem(
            source,
            client,
            output_path=destination,
            sheet_name="Data",
            identifier_column="Name",
            include_provenance=False,
            checkpoint_every=1,
            progress=_silent,
        )

    assert client.calls == 0
    assert partial.read_bytes() == b"user-owned checkpoint sentinel"
    assert not destination.exists()


def test_merged_table_body_and_column_limit_fail_before_provider(tmp_path: Path):
    merged_source = tmp_path / "merged.xlsx"
    _create_preservation_workbook(merged_source)
    workbook = openpyxl.load_workbook(merged_source)
    data = workbook["Data"]
    data["D1"] = "Resolved CAS"
    data.merge_cells("C2:D2")
    workbook.save(merged_source)
    workbook.close()
    merged_client = NoLookupPubChem()

    with pytest.raises(ValueError, match="merged range C2:D2"):
        run_pubchem(
            merged_source,
            merged_client,
            output_path=tmp_path / "merged-output.xlsx",
            sheet_name="Data",
            identifier_column="Name",
            include_provenance=False,
            checkpoint_every=0,
            progress=_silent,
        )
    assert merged_client.calls == 0

    wide_source = tmp_path / "wide.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Name"])
    worksheet.append(["C6"])
    assert {"query", "pubchem_url"}.issubset(PUBCHEM_COLUMN_MAP)
    no_odor_output_count = sum(key not in PUBCHEM_ODOR_KEYS for key in PUBCHEM_COLUMN_MAP) + 3
    first_unsafe_column = excel_io.MAX_EXCEL_COLUMNS - no_odor_output_count + 1
    worksheet.cell(row=1, column=first_unsafe_column, value="Reserved edge cell")
    workbook.save(wide_source)
    workbook.close()
    wide_client = NoLookupPubChem()

    with pytest.raises(ValueError, match="column limit"):
        run_pubchem(
            wide_source,
            wide_client,
            output_path=tmp_path / "wide-output.xlsx",
            identifier_column="Name",
            include_odor=False,
            include_provenance=False,
            checkpoint_every=0,
            progress=_silent,
        )
    assert wide_client.calls == 0


def test_custom_xml_part_is_rejected_before_provider(tmp_path: Path):
    source = tmp_path / "custom-xml.xlsx"
    _create_preservation_workbook(source)
    content_types = _xlsx_part(source, "[Content_Types].xml").replace(
        b"</Types>",
        (b'<Override PartName="/customXml/item1.xml" ContentType="application/xml"/></Types>'),
    )
    root_relationships = _xlsx_part(source, "_rels/.rels").replace(
        b"</Relationships>",
        (
            b'<Relationship Id="rIdAromaNexusCustom" '
            b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            b'relationships/customXml" Target="customXml/item1.xml"/>'
            b"</Relationships>"
        ),
    )
    _rewrite_xlsx(
        source,
        {
            "[Content_Types].xml": content_types,
            "_rels/.rels": root_relationships,
        },
        {"customXml/item1.xml": b'<?xml version="1.0"?><aroma>preserve me</aroma>'},
    )
    client = NoLookupPubChem()

    with pytest.raises(ValueError, match="customXml/item1.xml"):
        run_pubchem(
            source,
            client,
            output_path=tmp_path / "custom-output.xlsx",
            sheet_name="Data",
            identifier_column="Name",
            checkpoint_every=0,
            progress=_silent,
        )
    assert client.calls == 0


def test_same_file_alias_is_rejected_before_provider(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    alias = tmp_path / "hardlink.xlsx"
    _create_preservation_workbook(source)
    try:
        os.link(source, alias)
    except OSError as exc:
        pytest.skip(f"Hard links are unavailable: {exc}")
    client = NoLookupPubChem()

    with pytest.raises(ValueError, match="Input and output paths must differ"):
        run_pubchem(
            source,
            client,
            output_path=alias,
            sheet_name="Data",
            identifier_column="Name",
            force=True,
            checkpoint_every=0,
            progress=_silent,
        )
    assert client.calls == 0


def test_atomic_no_replace_wins_destination_race(tmp_path: Path, monkeypatch):
    source = tmp_path / "input.xlsx"
    destination = tmp_path / "output.xlsx"
    _create_preservation_workbook(source)
    frame, context = excel_io.read_table_with_context(source, sheet_name="Data")
    original_writer = excel_io._write_preserved_xlsx

    def racing_writer(frame, temporary, table_context):
        original_writer(frame, temporary, table_context)
        destination.write_bytes(b"competing writer")

    monkeypatch.setattr(excel_io, "_write_preserved_xlsx", racing_writer)
    with pytest.raises(FileExistsError, match="Pass --force"):
        excel_io.write_table(frame, destination, context=context)

    assert destination.read_bytes() == b"competing writer"
    assert not list(tmp_path.glob(".output-*.xlsx"))


def test_keyboard_interrupt_cleans_temporary_file(tmp_path: Path, monkeypatch):
    source = tmp_path / "input.xlsx"
    destination = tmp_path / "output.xlsx"
    _create_preservation_workbook(source)
    frame, context = excel_io.read_table_with_context(source, sheet_name="Data")

    def interrupting_writer(*_args, **_kwargs):
        raise KeyboardInterrupt

    monkeypatch.setattr(excel_io, "_write_preserved_xlsx", interrupting_writer)
    with pytest.raises(KeyboardInterrupt):
        excel_io.write_table(frame, destination, context=context)

    assert not destination.exists()
    assert not list(tmp_path.glob(".output-*.xlsx"))


def test_public_read_table_does_not_retain_private_workbook_snapshot(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    _create_preservation_workbook(source)

    frame = excel_io.read_table(source, sheet_name="Data")

    assert frame.attrs == {}
