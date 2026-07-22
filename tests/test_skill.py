import json
import re
import subprocess
import sys
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import openpyxl
import pandas as pd
from openpyxl.styles import Font

REPO_ROOT = Path(__file__).resolve().parents[1]
SKILL_ROOT = REPO_ROOT / ".agents" / "skills" / "curate-aroma-data"


def test_skill_metadata_and_interface_are_complete():
    skill = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
    interface = (SKILL_ROOT / "agents" / "openai.yaml").read_text(encoding="utf-8")
    assert skill.startswith("---\nname: curate-aroma-data\n")
    assert "TODO" not in skill
    assert "description:" in skill.split("---", 2)[1]
    assert 'display_name: "Curate Aroma Data"' in interface
    assert "$curate-aroma-data" in interface


def test_skill_inspector_reports_identifier_quality(tmp_path: Path):
    workbook = tmp_path / "input.xlsx"
    pd.DataFrame({"CAS Number": ["100-52-7", "", "100-52-8"]}).to_excel(workbook, index=False)
    script = SKILL_ROOT / "scripts" / "inspect_workbook.py"
    completed = subprocess.run(
        [sys.executable, str(script), str(workbook)],
        cwd=REPO_ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    report = json.loads(completed.stdout)
    assert report["rows"] == 3
    assert report["workbook"]["sheets"] == ["Sheet1"]
    assert report["workbook"]["selected_sheet"] == "Sheet1"
    assert report["cas"]["valid"] == 1
    assert report["cas"]["invalid"] == 1


def test_skill_inspector_reports_workbook_fidelity_evidence_and_clean_errors(tmp_path: Path):
    workbook_path = tmp_path / "input.xlsx"
    workbook = openpyxl.Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover["A1"] = "Read me"
    data = workbook.create_sheet("Data")
    data.append(["CAS Number", "Value"])
    data.append(["100-52-7", 10])
    data.append(["", "=B2*2"])
    data["A1"].font = Font(bold=True)
    data.column_dimensions["A"].width = 20
    data.row_dimensions[1].height = 24
    data.freeze_panes = "A2"
    data.auto_filter.ref = "A1:B3"
    data.merge_cells("A5:B5")
    data["A5"] = "Outside table"
    workbook.properties.title = "Inspector fixture"
    workbook.properties.creator = "AromaNexus tests"
    workbook.save(workbook_path)
    workbook.close()
    rewritten = BytesIO()
    with ZipFile(BytesIO(workbook_path.read_bytes())) as source, ZipFile(rewritten, "w") as output:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet2.xml":
                payload, replaced = re.subn(
                    rb"<v\s*(?:/\s*>|>\s*</v>)",
                    b"<v>20</v>",
                    payload,
                    count=1,
                )
                assert replaced == 1
            output.writestr(item, payload)
    workbook_path.write_bytes(rewritten.getvalue())
    script = SKILL_ROOT / "scripts" / "inspect_workbook.py"

    completed = subprocess.run(
        [sys.executable, str(script), str(workbook_path), "--sheet", "Data"],
        cwd=REPO_ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    report = json.loads(completed.stdout)
    workbook_report = report["workbook"]
    assert workbook_report["title"] == "Inspector fixture"
    assert workbook_report["creator"] == "AromaNexus tests"
    assert workbook_report["selected_sheet"] == "Data"
    assert set(workbook_report["sheet_summaries"]) == {"Cover", "Data"}
    selected = workbook_report["selected_sheet_summary"]
    assert selected["formula_cells"] == 1
    assert selected["cached_formula_results"] == 1
    assert len(selected["formula_cache_sha256"]) == 64
    assert selected["styled_cells"] >= 1
    assert selected["row_heights"] == {"1": 24.0}
    assert selected["column_widths"] == {"A": 20.0}
    assert selected["freeze_panes"] == "A2"
    assert selected["auto_filter"] == "A1:B3"
    assert selected["merged_ranges"] == ["A5:B5"]
    assert len(selected["content_sha256"]) == 64

    missing = subprocess.run(
        [sys.executable, str(script), str(workbook_path), "--sheet", "Missing"],
        cwd=REPO_ROOT,
        check=False,
        capture_output=True,
        text=True,
    )
    assert missing.returncode == 2
    assert missing.stdout == ""
    assert missing.stderr.startswith("Error: Worksheet 'Missing' was not found")
    assert "Traceback" not in missing.stderr

    missing_cas = subprocess.run(
        [
            sys.executable,
            str(script),
            str(workbook_path),
            "--sheet",
            "Data",
            "--cas-column",
            "Missing CAS",
        ],
        cwd=REPO_ROOT,
        check=False,
        capture_output=True,
        text=True,
    )
    assert missing_cas.returncode == 2
    assert missing_cas.stdout == ""
    assert missing_cas.stderr.startswith("Error: CAS column 'Missing CAS' was not found")
    assert "Traceback" not in missing_cas.stderr
